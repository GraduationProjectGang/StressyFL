import json
import time
import pprint
from openpyxl import Workbook
from ast import literal_eval
from dataclasses import dataclass
import numpy as np
from datetime import datetime
import time
# from datamake_rotate import getRotateVec
import pprint

pp = pprint.PrettyPrinter(indent=4)

def scoreMatching():
    jsonPath = 'C:\\Users\\ksh04\\PythonProjects\\IfMoving\\datacollect.json'

    with open(jsonPath, encoding= 'UTF-8') as json_file:
        usersData = json.load(json_file)
        users = usersData['user']

        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'scoreMatching'
        
        sheet1.cell(row=1, column=1).value = "userName"
        sheet1.cell(row=1, column=2).value = "timestamp"
        sheet1.cell(row=1, column=3).value = "strCount"
        data={}
        rowNum = 2
        num = 1
        stressJsonList = []
        for userName in users:
                data[userName] = []
                for property in users[userName]:
                        if property == 'location':
                                for locID in users[userName][property]: #locationID
                                        locItem = users[userName][property][locID]
                                        latArr = [] #latitudeArray
                                        lonArr = [] #longitudeArray
                                        speedArr = [] 
                                        accuracyArr = [] 
                                        for item in locItem: #0,1,2,3
                                                if item == "locationList":
                                                        for locListItem in users[userName][property][locID]['locationList']:
                                                                latitude = locListItem['latitude']
                                                                longitude = locListItem['longitude']
                                                                speed = locListItem['speed']
                                                                accuracy = locListItem['accuracy']
                                                                latArr.append(latitude)
                                                                lonArr.append(longitude)
                                                                speedArr.append(speed)
                                                                accuracyArr.append(accuracy)
                                                        # print(f'lat: {latitude} lon:{longitude} speeD:{speed}')  
                                                if item == "timestmamp":
                                                        dateTime = users[userName][property][locID]['timestmamp']
                                                        timestamp = time.mktime(datetime.strptime(dateTime, '%Y%m%d.%H:%M:%S').timetuple())
                                                        # print(timestamp)
                                                # print(len(latArr), len(lonArr))
                                                latAverage = 0 if len(latArr) == 0 else np.mean(latArr)
                                                lonAverage = 0 if len(lonArr) == 0 else np.mean(lonArr)
                                                speedMax = 0 if len(speedArr) == 0 else np.max(speedArr)
                                                accuracyAverage = 0 if len(accuracyArr) == 0 else np.mean(accuracyArr)
                                        # print(latAverage,lonAverage)


                                        sheet1.cell(row=rowNum, column=1).value = userName
                                        sheet1.cell(row=rowNum, column=2).value = dateTime
                                        sheet1.cell(row=rowNum, column=3).value = timestamp
                                        sheet1.cell(row=rowNum, column=4).value = latAverage
                                        sheet1.cell(row=rowNum, column=5).value = lonAverage
                                        sheet1.cell(row=rowNum, column=6).value = speedMax
                                        sheet1.cell(row=rowNum, column=7).value = accuracyAverage
                                        sheet1.cell(row=rowNum, column=8).value = 1 if speedMax > 1 else 0
                                        rowNum = rowNum + 1
                                        ifMoving = 1 if speedMax > 1 else 0
                                        data[userName].append({
                        "user":userName,
                        "timestamp":timestamp,
                        "ifMoving":ifMoving
                        })

                                        # pp.pprint(data[userName])
                        # print(data[userName])

                        # if property == 'stress':
                        #         for stressID in users[userName][property]: 
                        #                 stressItem = users[userName][property][stressID]
                        #                 for item in stressItem:
                        #                         if item == "timestamp":
                        #                                 timestamp = int(int(users[userName][property][stressID][item])/1000)
                        #                                 print(timestamp)
                        #                         if item == "stressCount":
                        #                                 stressCount = users[userName][property][stressID][item]
                        #                                 print(stressCount)            


                

                        
        count = 1
        # for userKey in data:
        #         for attr in data[userKey]:  
        #                 for rotateUserKey in rotateData:
        #                         for attr_rotate in rotateData[rotateUserKey]:
        #                                 if userKey == rotateUserKey:
        #                                         if  attr['timestamp'] == attr_rotate['timestamp']:
                                                        
        #                                                 for iter in attr_rotate:
        #                                                         if iter == 'posture' or iter == 'posture_accuracy' or iter == 'std_posture' or iter == 'orientation' or iter == 'stressCount':
        #                                                                 attr[iter] = attr_rotate[iter]

                                                                        
        # statspath = 'C:\\Users\\ksh04\\PythonProjects\\DataManufacture\\appstats.json'

        # with open(statspath, encoding= 'UTF-8') as file:
        #         statsData = json.load(file)

        #         for userKey in data:
        #                 for item in data[userKey]: #jsonItem - user,timestamp,ifMoving,posture,posture_accuracy,std_posture,orientation
        #                         for user in statsData:
        #                                 if userKey == user:
        #                                         for coroutine in statsData[userKey]:
        #                                                 if item['timestamp'] == statsData[userKey][coroutine]['timestamp']:
        #                                                         for apps in statsData[userKey][coroutine]:
        #                                                                 if len(apps) == 1:
        #                                                                         data[userKey][item][apps] = statsData[userKey][coroutine][apps]
                                                                
                                                


                                


        wb.save(filename= 'data2.xlsx')

        return data




        

if __name__ == "__main__":
        data = scoreMatching()
        with open('data.json', 'w') as outfile:
                json.dump(data, outfile)
        

        pprint.pprint(data['-MAfh2YZ9fJ8BtZKBp7K'])
